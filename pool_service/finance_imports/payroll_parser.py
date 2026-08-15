from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import re

from openpyxl import load_workbook

from .employee_matching import normalize_onec_name
from .validators import validate_xlsx_archive


PARSER_VERSION = "1"
MAX_ROWS = 100_000
CONTROL_TOLERANCE = Decimal("0.01")


class PayrollParseError(ValueError):
    pass


@dataclass
class PayrollParseResult:
    records: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    critical_errors: list[str] = field(default_factory=list)
    metadata: dict = field(default_factory=dict)


def _text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def _decimal(value):
    if value in (None, "", "-", "—"):
        return Decimal("0.00")
    if isinstance(value, bool):
        raise PayrollParseError("Логическое значение не является суммой.")
    try:
        return Decimal(str(value).replace(" ", "").replace(",", ".")).quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise PayrollParseError(f"Не удалось распознать сумму: {value!r}") from exc


def _month(value):
    if isinstance(value, (date, datetime)):
        return date(value.year, value.month, 1)
    text = _text(value).casefold().replace("ё", "е")
    match = re.search(r"\b(20\d{2})[-/.](0?[1-9]|1[0-2])\b", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)
    match = re.search(r"\b(0?[1-9]|1[0-2])[./-](20\d{2})\b", text)
    if match:
        return date(int(match.group(2)), int(match.group(1)), 1)
    return None


def _totals(values):
    return tuple(_decimal(value) for value in values)


def _different(left, right):
    return any(abs(a - b) > CONTROL_TOLERANCE for a, b in zip(left, right))


def parse_payroll(file_obj, *, filename=None, size=None):
    validate_xlsx_archive(file_obj, filename=filename, size=size)
    workbook = load_workbook(file_obj, read_only=False, data_only=True)
    if len(workbook.worksheets) != 1:
        raise PayrollParseError("Отчёт по расчётам с персоналом должен содержать один лист.")
    sheet = workbook.worksheets[0]
    if sheet.max_row > MAX_ROWS or sheet.max_column != 6:
        raise PayrollParseError("Неожиданная размерность отчёта по расчётам с персоналом.")

    result = PayrollParseResult()
    department = ""
    department_control = None
    department_records = []
    employee = ""
    personnel_number = ""
    employee_control = None
    employee_records = []

    def validate_employee():
        if employee and employee_control is not None:
            actual = tuple(sum((row[key] for row in employee_records), Decimal("0")) for key in (
                "opening_balance", "accrued", "paid", "closing_balance"
            ))
            if _different(employee_control, actual):
                result.critical_errors.append(f"Контрольные итоги сотрудника не совпадают: {employee}.")

    def validate_department():
        if department and department_control is not None:
            actual = tuple(sum((row[key] for row in department_records), Decimal("0")) for key in (
                "opening_balance", "accrued", "paid", "closing_balance"
            ))
            if _different(department_control, actual):
                result.critical_errors.append(f"Контрольные итоги подразделения не совпадают: {department}.")

    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_col=6), 1):
        label = _text(row[0].value)
        if not label:
            continue
        indent = int(row[0].alignment.indent or 0)
        values = [cell.value for cell in row[2:6]]
        if indent == 0:
            validate_employee(); validate_department()
            department, department_control = label, _totals(values)
            department_records = []
            employee = ""; employee_records = []
        elif indent == 1:
            validate_employee()
            if not department:
                raise PayrollParseError(f"Сотрудник вне подразделения, строка {row_number}.")
            employee, personnel_number = label, _text(row[1].value)
            employee_control = _totals(values)
            employee_records = []
        elif indent == 2:
            period = _month(label)
            if not department or not employee or period is None:
                raise PayrollParseError(f"Некорректная строка периода, строка {row_number}.")
            amounts = _totals(values)
            record = {
                "period_month": period,
                "source_row_number": row_number,
                "department_name": department,
                "employee_raw_name": employee,
                "employee_normalized_name": normalize_onec_name(employee),
                "personnel_number": personnel_number,
                "opening_balance": amounts[0],
                "accrued": amounts[1],
                "paid": amounts[2],
                "closing_balance": amounts[3],
                "source_data": {"hierarchy_level": "period"},
            }
            result.records.append(record)
            employee_records.append(record)
            department_records.append(record)
        elif any(value not in (None, "") for value in values):
            raise PayrollParseError(f"Неизвестный уровень иерархии, строка {row_number}.")
    validate_employee(); validate_department()
    if not result.records:
        raise PayrollParseError("В отчёте не найдены строки Employee/Period.")
    periods = sorted({row["period_month"] for row in result.records})
    result.metadata = {
        "format": "department_employee_period_v1",
        "period_first": periods[0].isoformat(),
        "period_last": periods[-1].isoformat(),
        "row_count": len(result.records),
        "control_totals": {
            key: str(sum((row[key] for row in result.records), Decimal("0")))
            for key in ("opening_balance", "accrued", "paid", "closing_balance")
        },
    }
    return result
