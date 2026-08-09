from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .validators import validate_xlsx_archive

PARSER_VERSION = "2"
FORMAT_HORIZONTAL = "horizontal"
FORMAT_VERTICAL_1C = "vertical_1c"
MAX_SHEETS = 10
MAX_ROWS = 100_000
MAX_COLUMNS = 500
HEADER_SCAN_ROWS = 120
MAX_TOTAL_CELLS = 5_000_000
PARSER_WARNING_BUFFER = 200
MAX_SOURCE_CELLS = 50
MAX_SOURCE_VALUE_LENGTH = 500
MAX_FORMULA_CELLS = 10_000

MONTHS = {
    "январь": 1, "янв": 1, "февраль": 2, "февр": 2, "фев": 2, "март": 3, "мар": 3,
    "апрель": 4, "апр": 4, "май": 5, "июнь": 6, "июн": 6, "июль": 7,
    "июл": 7, "август": 8, "авг": 8, "сентябрь": 9, "сент": 9, "сен": 9,
    "октябрь": 10, "окт": 10, "ноябрь": 11, "нояб": 11, "ноя": 11,
    "декабрь": 12, "дек": 12,
}
METRICS = {
    "profitability_percent": ("% прибыли", "процент прибыли", "рентабельность", "рентаб"),
    "quantity": ("количество", "кол-во", "кол во"),
    "revenue": ("выручка", "продажа", "сумма продаж"),
    "cost": ("себестоимость", "себест"),
    "gross_profit": ("сумма валовой прибыли", "валовая прибыль"),
}
NAME_MARKERS = ("номенклатура", "наименование", "товар", "продукция")
ARTICLE_MARKERS = ("артикул", "код")
TOTAL_MARKERS = ("итого", "всего", "общий итог")
TOTAL_TOLERANCE = Decimal("0.05")


class MonthlyProfitParseError(ValueError):
    pass


@dataclass
class ParseResult:
    records: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    warnings_total: int = 0
    critical_errors: list[str] = field(default_factory=list)
    metadata: dict = field(default_factory=dict)
    rows_read: int = 0
    rows_skipped: int = 0

    def add_warning(self, message):
        self.warnings_total += 1
        if len(self.warnings) < PARSER_WARNING_BUFFER:
            self.warnings.append(str(message))


def _text(value):
    return "" if value is None else re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def parse_decimal(value, *, percent=False):
    del percent  # Percent values are stored as percentage points, not fractions.
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ValueError("Логическое значение не является числом.")
    if isinstance(value, (int, Decimal)):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    raw = _text(value)
    if not raw or raw in {"-", "—", "–"}:
        return None
    negative = raw.startswith("(") and raw.endswith(")")
    if negative:
        raw = raw[1:-1].strip()
    raw = raw.replace("%", "").replace(" ", "").replace("'", "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".") if raw.rfind(",") > raw.rfind(".") else raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".") if raw.count(",") == 1 else raw.replace(",", "")
    try:
        result = Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Не удалось распознать число: {value!r}") from exc
    return -result if negative else result


DECIMAL_RULES = {
    "quantity": (Decimal("0.000001"), Decimal("1e14")),
    "revenue": (Decimal("0.01"), Decimal("1e18")),
    "cost": (Decimal("0.01"), Decimal("1e18")),
    "gross_profit": (Decimal("0.01"), Decimal("1e18")),
    "profitability_percent": (Decimal("0.0001"), Decimal("1e8")),
}


def normalize_metric_decimal(value, metric):
    if value is None:
        return None
    quantum, limit = DECIMAL_RULES[metric]
    if abs(value) >= limit:
        raise ValueError("Число превышает допустимую точность.")
    try:
        return value.quantize(quantum, rounding=ROUND_HALF_UP)
    except InvalidOperation as exc:
        raise ValueError("Число превышает допустимую точность.") from exc


def parse_month(value, *, fallback_year=None, workbook_epoch=None):
    if isinstance(value, (date, datetime)):
        return date(value.year, value.month, 1)
    if isinstance(value, (int, float)) and workbook_epoch is not None:
        try:
            converted = from_excel(value, workbook_epoch)
            if isinstance(converted, (date, datetime)) and 2000 <= converted.year <= 2100:
                return date(converted.year, converted.month, 1)
        except (TypeError, ValueError, OverflowError):
            pass
    normalized = _text(value).lower().replace("ё", "е")
    match = re.search(r"\b(20\d{2})[-/.](0?[1-9]|1[0-2])\b", normalized)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)
    match = re.search(r"\b(0?[1-9]|1[0-2])[./-](20\d{2})\b", normalized)
    if match:
        return date(int(match.group(2)), int(match.group(1)), 1)
    month_text = re.sub(r"[.]", "", normalized)
    for name, month in sorted(MONTHS.items(), key=lambda item: -len(item[0])):
        if re.search(rf"\b{re.escape(name)}\b", month_text):
            year_match = re.search(r"\b(20\d{2})\b", month_text)
            year = int(year_match.group(0)) if year_match else fallback_year
            return date(year, month, 1) if year else (None, month)
    return None


def _metric(text):
    normalized = _text(text).lower().replace("ё", "е")
    for key, markers in METRICS.items():
        if any(marker in normalized for marker in markers):
            return key
    return None


def _json_value(value):
    if value is None or isinstance(value, (int, bool)):
        return value
    if isinstance(value, str):
        return value[:MAX_SOURCE_VALUE_LENGTH]
    if isinstance(value, (float, Decimal)):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)[:MAX_SOURCE_VALUE_LENGTH]


def _detect_report_year(rows):
    years = []
    for row in rows[:60]:
        years.extend(int(item) for item in re.findall(r"\b20\d{2}\b", " ".join(_text(v) for v in row)))
    return years[0] if years and len(set(years)) == 1 else None


def _find_header(rows, year, epoch):
    best = None
    for start in range(min(len(rows), HEADER_SCAN_ROWS)):
        for depth in (1, 2, 3):
            block = rows[start:start + depth]
            width = max((len(row) for row in block), default=0)
            columns, name_col, article_col = [], None, None
            for col in range(width):
                pieces = [_text(row[col]) if col < len(row) else "" for row in block]
                combined = " ".join(filter(None, pieces))
                lowered = combined.lower().replace("ё", "е")
                if name_col is None and any(marker in lowered for marker in NAME_MARKERS): name_col = col
                if article_col is None and any(marker in lowered for marker in ARTICLE_MARKERS): article_col = col
                month = next((parse_month(piece, fallback_year=year, workbook_epoch=epoch) for piece in pieces if parse_month(piece, fallback_year=year, workbook_epoch=epoch)), None)
                columns.append({"index": col, "month": month, "metric": _metric(combined)})
            last_month = None
            for column in columns:
                if column["month"]: last_month = column["month"]
                elif last_month and column["metric"]: column["month"] = last_month
            metric_columns = [column for column in columns if column["metric"] and column["month"]]
            score = len(metric_columns) * 10 + (5 if name_col is not None else 0)
            if metric_columns and (best is None or score > best[0]):
                best = (score, start, depth, columns, name_col, article_col)
    if not best:
        raise MonthlyProfitParseError("Не удалось найти заголовки месяцев и показателей.")
    _, start, depth, columns, name_col, article_col = best
    if name_col is None:
        used = {column["index"] for column in columns if column["metric"]}
        name_col = next((index for index in range(max(used or {0}) + 1) if index not in used), 0)
    return start, depth, columns, name_col, article_col


def _is_total_label(value):
    normalized = _text(value).lower().replace("ё", "е")
    return bool(re.match(r"^(?:общий\s+итог|итого|всего)(?:\s*$|\s+(?:по|за)\b|\s*:)", normalized))


def _parse_horizontal_rows(sheet_title, rows, epoch, formula_cells=None, formulas_truncated=False):
    year = _detect_report_year(rows)
    start, depth, columns, name_col, article_col = _find_header(rows, year, epoch)
    result = ParseResult()
    unresolved = [c for c in columns if c["metric"] and isinstance(c["month"], tuple)]
    if unresolved: result.critical_errors.append("Найден месяц без года, а год отчёта определить не удалось.")
    metric_columns = [c for c in columns if c["metric"] and isinstance(c["month"], date)]
    months = sorted({c["month"] for c in metric_columns})
    for excel_row, row in enumerate(rows[start + depth:], start + depth + 1):
        result.rows_read += 1
        name = _text(row[name_col] if name_col < len(row) else "")
        if not name or _is_total_label(name):
            result.rows_skipped += 1; continue
        article = _text(row[article_col] if article_col is not None and article_col < len(row) else "")
        by_month = {}
        for column in metric_columns:
            raw = row[column["index"]] if column["index"] < len(row) else None
            try:
                value = normalize_metric_decimal(
                    parse_decimal(raw, percent=column["metric"] == "profitability_percent"),
                    column["metric"],
                )
            except ValueError:
                value = None
                result.add_warning(f"Строка {excel_row}: неверное число в показателе {column['metric']}.")
            by_month.setdefault(column["month"], {})[column["metric"]] = value
        if not any(any(v is not None for v in values.values()) for values in by_month.values()):
            result.rows_skipped += 1; continue
        source = {
            str(i + 1): _json_value(v)
            for i, v in enumerate(row[:MAX_SOURCE_CELLS])
            if v is not None
        }
        for month, values in sorted(by_month.items()):
            if not any(v is not None for v in values.values()): continue
            result.records.append({
                "period_month": month, "source_row_number": excel_row,
                "nomenclature": name[:500], "article": article[:120],
                "quantity": values.get("quantity"), "revenue": values.get("revenue"),
                "cost": values.get("cost"), "gross_profit": values.get("gross_profit"),
                "profitability_percent": values.get("profitability_percent"),
                "source_data": {
                    "excel_row_number": excel_row,
                    "detected_layout": FORMAT_HORIZONTAL,
                    "cells": source,
                },
            })
    for row_number, column_number in formula_cells or ():
        cached_value = (
            rows[row_number - 1][column_number - 1]
            if row_number <= len(rows) and column_number <= len(rows[row_number - 1])
            else None
        )
        if cached_value is None:
            result.add_warning(
                f"Ячейка R{row_number}C{column_number} содержит формулу без сохранённого значения."
            )
    if formulas_truncated:
        result.add_warning("Часть формул не перечислена из-за лимита диагностических ячеек.")
    if not result.records:
        result.critical_errors.append("В отчёте не обнаружено строк с числовыми показателями.")
    result.metadata = {
        "sheet": sheet_title, "layout": FORMAT_HORIZONTAL,
        "header_row": start + 1, "header_depth": depth,
        "report_year": year, "months": [m.isoformat() for m in months], "month_count": len(months),
        "rows_read": result.rows_read, "rows_skipped": result.rows_skipped,
    }
    return result


def _source_cells(row):
    return {
        str(index + 1): _json_value(value)
        for index, value in enumerate(row[:MAX_SOURCE_CELLS])
        if value is not None
    }


def _add_formula_warnings(result, rows, formula_cells, formulas_truncated):
    for row_number, column_number in formula_cells or ():
        cached_value = (
            rows[row_number - 1][column_number - 1]
            if row_number <= len(rows) and column_number <= len(rows[row_number - 1])
            else None
        )
        if cached_value is None:
            result.add_warning(
                f"Ячейка R{row_number}C{column_number} содержит формулу без сохранённого значения."
            )
    if formulas_truncated:
        result.add_warning("Часть формул не перечислена из-за лимита диагностических ячеек.")


def _find_vertical_header(rows, epoch):
    report_year = _detect_report_year(rows)
    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        first = _text(row[0] if row else "").lower().replace("ё", "е")
        metrics = {}
        article_col = None
        for column, value in enumerate(row):
            text = _text(value).lower().replace("ё", "е")
            metric = _metric(text)
            if metric and metric not in metrics:
                metrics[metric] = column
            if article_col is None and any(marker in text for marker in ARTICLE_MARKERS):
                article_col = column
        financial = {"revenue", "cost", "gross_profit", "profitability_percent"}
        if first in {"месяц", "период"} and len(financial.intersection(metrics)) >= 3:
            has_month_rows = any(
                _parse_vertical_month(item[0] if item else None, epoch, report_year) is not None
                for item in rows[index + 1:]
            )
            if has_month_rows:
                return index, 0, article_col, metrics, report_year
    return None


def detect_layout(rows, epoch):
    return FORMAT_VERTICAL_1C if _find_vertical_header(rows, epoch) else FORMAT_HORIZONTAL


def _parse_vertical_month(value, epoch, fallback_year=None):
    if isinstance(value, (date, datetime)):
        return parse_month(value, workbook_epoch=epoch)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return parse_month(value, workbook_epoch=epoch)
    normalized = _text(value).lower().replace("ё", "е")
    if re.fullmatch(r"(?:20\d{2}[-/.](?:0?[1-9]|1[0-2])|(?:0?[1-9]|1[0-2])[./-]20\d{2})", normalized):
        return parse_month(value, workbook_epoch=epoch)
    parts = normalized.replace(".", "").split()
    if len(parts) == 2 and parts[0] in MONTHS and re.fullmatch(r"20\d{2}", parts[1]):
        return parse_month(value, workbook_epoch=epoch)
    if len(parts) == 1 and parts[0] in MONTHS:
        return parse_month(value, fallback_year=fallback_year, workbook_epoch=epoch)
    return None


def _detect_vertical_hierarchy(
    rows, row_indents, start, name_col, article_col, metric_columns, epoch, fallback_year,
):
    current_month = None
    candidates = []
    has_nomenclature_schema_marker = False
    for offset, row in enumerate(rows[start + 1:], start + 1):
        name = row[name_col] if name_col < len(row) else None
        month = _parse_vertical_month(name, epoch, fallback_year)
        if isinstance(month, date):
            current_month = month
            continue
        if isinstance(month, tuple):
            current_month = None
            continue
        normalized_name = _text(name).lower().replace("ё", "е")
        if current_month is None and normalized_name == "номенклатура":
            has_nomenclature_schema_marker = True
        if not current_month or not _text(name) or _is_total_label(name):
            continue
        has_values = any(
            column < len(row) and row[column] not in (None, "", "-", "—", "–")
            for column in metric_columns.values()
        )
        if has_values:
            article = _text(
                row[article_col]
                if article_col is not None and article_col < len(row)
                else ""
            )
            candidates.append((row_indents[offset], bool(article)))

    if candidates and article_col is not None and all(has_article for _, has_article in candidates):
        return {
            "status": "flat",
            "detail_indent": None,
            "reason": "explicit_flat_schema",
        }

    positive_indents = {indent for indent, _ in candidates if indent > 0}
    has_lower_financial_level = any(
        indent < next(iter(positive_indents)) for indent, _ in candidates
    ) if len(positive_indents) == 1 else False
    if len(positive_indents) == 1 and (
        has_lower_financial_level or has_nomenclature_schema_marker
    ):
        return {
            "status": "reliable",
            "detail_indent": next(iter(positive_indents)),
            "reason": "stable_positive_indent",
        }
    if len(positive_indents) > 1:
        reason = "conflicting_detail_levels"
    else:
        reason = "no_reliable_detail_level"
    return {"status": "ambiguous", "detail_indent": None, "reason": reason}


def _serialized_totals(values):
    return {
        key: value if isinstance(value, int) else str(value) if value is not None else None
        for key, value in values.items()
    }


def _parse_vertical_rows(
    sheet_title, rows, row_indents, epoch, header, formula_cells=None, formulas_truncated=False,
):
    start, name_col, article_col, metric_columns, report_year = header
    result = ParseResult()
    current_month = None
    months = []
    source_totals = None
    monetary = ("revenue", "cost", "gross_profit")
    calculated = {key: Decimal("0") for key in monetary}
    month_totals = {}
    hierarchy = _detect_vertical_hierarchy(
        rows, row_indents, start, name_col, article_col, metric_columns, epoch, report_year
    )
    detail_indent = hierarchy["detail_indent"]
    hierarchy_status = hierarchy["status"]
    hierarchy_reason = hierarchy["reason"]
    aggregate_rows_skipped = 0

    for offset, row in enumerate(rows[start + 1:], start + 1):
        excel_row = offset + 1
        result.rows_read += 1
        name = _text(row[name_col] if name_col < len(row) else "")
        month = _parse_vertical_month(name, epoch, report_year)
        if isinstance(month, date):
            current_month = month
            if month not in months:
                months.append(month)
            result.rows_skipped += 1
            continue
        if isinstance(month, tuple):
            current_month = None
            result.critical_errors.append(
                f"Строка {excel_row}: найден месяц без года."
            )
            result.rows_skipped += 1
            continue
        if _is_total_label(name):
            parsed_totals = {}
            for metric in monetary:
                column = metric_columns.get(metric)
                if column is None:
                    continue
                raw = row[column] if column < len(row) else None
                try:
                    parsed_totals[metric] = normalize_metric_decimal(parse_decimal(raw), metric)
                except ValueError:
                    parsed_totals[metric] = None
                    result.add_warning(
                        f"Строка {excel_row}: неверное число в общем итоге {metric}."
                    )
            if parsed_totals:
                source_totals = parsed_totals
            result.rows_skipped += 1
            continue
        normalized_name = name.lower().replace("ё", "е")
        if not name or normalized_name in {"месяц", "номенклатура", "наименование"}:
            result.rows_skipped += 1
            continue
        if current_month is None:
            result.rows_skipped += 1
            continue
        if detail_indent is not None and row_indents[offset] != detail_indent:
            result.rows_skipped += 1
            aggregate_rows_skipped += 1
            result.add_warning(
                f"Строка {excel_row}: агрегированная строка пропущена для защиты от двойного учёта."
            )
            continue

        values = {}
        for metric, column in metric_columns.items():
            raw = row[column] if column < len(row) else None
            try:
                values[metric] = normalize_metric_decimal(
                    parse_decimal(raw, percent=metric == "profitability_percent"), metric
                )
            except ValueError:
                values[metric] = None
                result.add_warning(
                    f"Строка {excel_row}: неверное число в показателе {metric}."
                )
        if not any(values.get(metric) is not None for metric in monetary + ("profitability_percent",)):
            result.rows_skipped += 1
            continue

        record = {
            "period_month": current_month,
            "source_row_number": excel_row,
            "nomenclature": name[:500],
            "article": _text(
                row[article_col] if article_col is not None and article_col < len(row) else ""
            )[:120],
            "quantity": values.get("quantity"),
            "revenue": values.get("revenue"),
            "cost": values.get("cost"),
            "gross_profit": values.get("gross_profit"),
            "profitability_percent": values.get("profitability_percent"),
            "source_data": {
                "excel_row_number": excel_row,
                "detected_layout": FORMAT_VERTICAL_1C,
                "cells": _source_cells(row),
            },
        }
        result.records.append(record)
        monthly = month_totals.setdefault(
            current_month, {"rows": 0, **{key: Decimal("0") for key in monetary}}
        )
        monthly["rows"] += 1
        for metric in monetary:
            value = record[metric]
            if value is not None:
                calculated[metric] += value
                monthly[metric] += value

    totals_match = None
    if source_totals and any(source_totals.get(key) is not None for key in monetary):
        compared = [
            abs(calculated[key] - source_totals[key]) <= TOTAL_TOLERANCE
            for key in monetary if source_totals.get(key) is not None
        ]
        totals_match = bool(compared) and all(compared)
        if not totals_match:
            message = "Итоги распознанных строк существенно отличаются от общего итога 1С."
            result.add_warning(message)
            result.critical_errors.append(message)

    _add_formula_warnings(result, rows, formula_cells, formulas_truncated)
    if not months:
        result.critical_errors.append("В вертикальном отчёте не обнаружены строки месяцев.")
    elif not result.records:
        result.critical_errors.append(
            "Не найдено ни одной строки номенклатуры с финансовыми данными."
        )
    elif hierarchy_status == "ambiguous":
        result.critical_errors.append(
            "Не удалось надёжно определить уровень детальной номенклатуры; "
            "подтверждение заблокировано для защиты от двойного учёта."
        )

    years = {month.year for month in months}
    result.metadata = {
        "sheet": sheet_title,
        "layout": FORMAT_VERTICAL_1C,
        "header_row": start + 1,
        "header_depth": 1,
        "report_year": next(iter(years)) if len(years) == 1 else None,
        "months": [month.isoformat() for month in months],
        "month_count": len(months),
        "rows_read": result.rows_read,
        "rows_skipped": result.rows_skipped,
        "source_totals": _serialized_totals(source_totals or {}),
        "calculated_totals": _serialized_totals(calculated),
        "totals_match": totals_match,
        "hierarchy_status": hierarchy_status,
        "hierarchy_reason": hierarchy_reason,
        "aggregate_rows_skipped": aggregate_rows_skipped,
        "month_totals": {
            month.isoformat(): _serialized_totals(values)
            for month, values in month_totals.items()
        },
        "detail_indent": detail_indent,
    }
    return result


def _parse_sheet(sheet, epoch, formula_cells=None, formulas_truncated=False):
    if sheet.max_column and sheet.max_column > MAX_COLUMNS:
        raise MonthlyProfitParseError(f"Лист «{sheet.title}» содержит более {MAX_COLUMNS} колонок.")
    rows = []
    row_indents = []
    for number, cells in enumerate(sheet.iter_rows(), 1):
        if number > MAX_ROWS:
            raise MonthlyProfitParseError("Превышен лимит строк.")
        limited = cells[:MAX_COLUMNS]
        rows.append(tuple(cell.value for cell in limited))
        first = limited[0] if limited else None
        row_indents.append(float(getattr(getattr(first, "alignment", None), "indent", 0) or 0))
    layout = detect_layout(rows, epoch)
    if layout == FORMAT_VERTICAL_1C:
        return _parse_vertical_rows(
            sheet.title, rows, row_indents, epoch, _find_vertical_header(rows, epoch),
            formula_cells, formulas_truncated,
        )
    return _parse_horizontal_rows(
        sheet.title, rows, epoch, formula_cells, formulas_truncated,
    )


def parse_monthly_profit(file_obj, *, filename="report.xlsx", size=None):
    validate_xlsx_archive(file_obj, size=size, filename=filename)
    file_obj.seek(0)
    try:
        formula_workbook = load_workbook(file_obj, read_only=True, data_only=False, keep_links=False)
    except Exception as exc:
        raise MonthlyProfitParseError("Не удалось безопасно открыть XLSX.") from exc
    try:
        if len(formula_workbook.sheetnames) > MAX_SHEETS:
            raise MonthlyProfitParseError(f"Количество листов превышает лимит {MAX_SHEETS}.")
        declared_rows = sum(min(sheet.max_row or 0, MAX_ROWS + 1) for sheet in formula_workbook.worksheets)
        if declared_rows > MAX_ROWS:
            raise MonthlyProfitParseError(f"Суммарное количество строк превышает лимит {MAX_ROWS}.")
        declared_cells = sum((sheet.max_row or 0) * (sheet.max_column or 0) for sheet in formula_workbook.worksheets)
        if declared_cells > MAX_TOTAL_CELLS:
            raise MonthlyProfitParseError("Количество ячеек превышает безопасный лимит.")
        formula_cells, formula_truncated = {}, {}
        for sheet in formula_workbook.worksheets:
            cells, truncated = set(), False
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type != "f":
                        continue
                    if len(cells) < MAX_FORMULA_CELLS:
                        cells.add((cell.row, cell.column))
                    else:
                        truncated = True
            formula_cells[sheet.title] = cells
            formula_truncated[sheet.title] = truncated
    finally:
        formula_workbook.close()
        file_obj.seek(0)
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise MonthlyProfitParseError("Не удалось безопасно открыть XLSX.") from exc
    try:
        candidates, errors = [], []
        for sheet in workbook.worksheets:
            try:
                candidates.append(_parse_sheet(
                    sheet,
                    workbook.epoch,
                    formula_cells.get(sheet.title),
                    formula_truncated.get(sheet.title, False),
                ))
            except MonthlyProfitParseError as exc: errors.append(f"{sheet.title}: {exc}")
        if not candidates:
            raise MonthlyProfitParseError("Подходящий лист не найден. " + "; ".join(errors[:3]))
        result = max(candidates, key=lambda item: len(item.records))
        result.metadata["detected_sheets"] = list(workbook.sheetnames)
        return result
    finally:
        workbook.close()
        file_obj.seek(0)
