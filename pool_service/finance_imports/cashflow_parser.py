from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import re

from openpyxl import load_workbook

from .employee_matching import normalize_onec_name
from .validators import validate_xlsx_archive


PARSER_VERSION = "1"
MAX_ROWS = 100_000
CONTROL_TOLERANCE = Decimal("0.01")


class CashFlowParseError(ValueError):
    pass


@dataclass
class CashFlowParseResult:
    records: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    critical_errors: list[str] = field(default_factory=list)
    metadata: dict = field(default_factory=dict)


def _text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def _decimal(value):
    if value in (None, "", "-", "—"):
        return Decimal("0.00")
    if isinstance(value, bool):
        raise CashFlowParseError("Логическое значение не является суммой.")
    try:
        return Decimal(str(value).replace(" ", "").replace(",", ".")).quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise CashFlowParseError(f"Не удалось распознать сумму: {value!r}") from exc


def _month(value):
    if isinstance(value, (date, datetime)):
        return date(value.year, value.month, 1)
    text = _text(value)
    match = re.search(r"\b(20\d{2})[-/.](0?[1-9]|1[0-2])\b", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1)
    match = re.search(r"\b(0?[1-9]|1[0-2])[./-](20\d{2})\b", text)
    if match:
        return date(int(match.group(2)), int(match.group(1)), 1)
    return None


def _amounts(cells):
    receipts, payments, net = (_decimal(cell.value) for cell in cells)
    if abs((receipts - payments) - net) > CONTROL_TOLERANCE:
        raise CashFlowParseError("Чистый денежный поток не равен поступлениям минус платежи.")
    return receipts, payments, net


def parse_cashflow(file_obj, *, filename=None, size=None):
    validate_xlsx_archive(file_obj, filename=filename, size=size)
    workbook = load_workbook(file_obj, read_only=False, data_only=True)
    if len(workbook.worksheets) != 1:
        raise CashFlowParseError("Отчёт ДДС должен содержать один лист.")
    sheet = workbook.worksheets[0]
    if sheet.max_row > MAX_ROWS or sheet.max_column != 4:
        raise CashFlowParseError("Неожиданная размерность отчёта ДДС.")
    result = CashFlowParseResult()
    article = ""
    article_control = None
    article_records = []
    period = None
    month_control = None
    month_records = []

    def totals(records):
        return tuple(sum((row[key] for row in records), Decimal("0")) for key in (
            "receipts", "payments", "net_cash_flow"
        ))

    def validate(control, records, label):
        if control is not None and any(
            abs(left - right) > CONTROL_TOLERANCE for left, right in zip(control, totals(records))
        ):
            result.critical_errors.append(f"Контрольные итоги не совпадают: {label}.")

    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_col=4), 1):
        label = _text(row[0].value)
        if not label:
            continue
        indent = int(row[0].alignment.indent or 0)
        if indent == 0:
            validate(month_control, month_records, f"месяц {period}")
            validate(article_control, article_records, f"статья {article}")
            article, article_control = label, _amounts(row[1:4])
            article_records = []
            period = None; month_control = None; month_records = []
        elif indent == 1:
            validate(month_control, month_records, f"месяц {period}")
            if not article:
                raise CashFlowParseError(f"Месяц вне статьи, строка {row_number}.")
            period = _month(label)
            if period is None:
                raise CashFlowParseError(f"Не удалось распознать месяц, строка {row_number}.")
            month_control = _amounts(row[1:4])
            month_records = []
        elif indent == 2:
            if not article or period is None:
                raise CashFlowParseError(f"Документ вне Article/Month, строка {row_number}.")
            receipts, payments, net = _amounts(row[1:4])
            record = {
                "period_month": period,
                "source_row_number": row_number,
                "source_reference": "",
                "article_raw": article,
                "normalized_article_name": normalize_onec_name(article),
                "document_raw": label,
                "receipts": receipts,
                "payments": payments,
                "net_cash_flow": net,
                "source_data": {"hierarchy_level": "document"},
            }
            result.records.append(record)
            month_records.append(record)
            article_records.append(record)
        else:
            raise CashFlowParseError(f"Неизвестный уровень иерархии, строка {row_number}.")
    validate(month_control, month_records, f"месяц {period}")
    validate(article_control, article_records, f"статья {article}")
    if not result.records:
        raise CashFlowParseError("В отчёте не найдены document rows.")
    periods = sorted({row["period_month"] for row in result.records})
    result.metadata = {
        "format": "article_month_document_v1",
        "period_first": periods[0].isoformat(),
        "period_last": periods[-1].isoformat(),
        "row_count": len(result.records),
        "control_totals": {
            key: str(sum((row[key] for row in result.records), Decimal("0")))
            for key in ("receipts", "payments", "net_cash_flow")
        },
    }
    return result
