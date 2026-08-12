from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
import zipfile

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, transaction
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment

from pool_service.finance_forms import MonthlyProfitUploadForm
from pool_service.finance_imports.monthly_profit_parser import (
    ParseResult,
    classify_nomenclature_type,
    parse_decimal,
    parse_monthly_profit,
)
from pool_service.finance_imports.services import (
    DuplicateImportError,
    _preview_metadata,
    calculate_profitability,
    cancel_monthly_profit,
    confirm_monthly_profit,
    create_monthly_profit_preview,
    apply_period_weighted_goods_cost,
)
from pool_service.finance_imports.validators import delete_private_batch_file
from pool_service.models import OneCImportBatch, OneCMonthlyProfit, Organization, OrganizationAccess


def xlsx_bytes(
    *, year=2026, month_label="Январь 2026", rows=None, multilevel=True,
    profitability_label="Рентабельность",
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Валовая прибыль"
    sheet.append([f"Валовая прибыль за {year}" if year else "Валовая прибыль"])
    if multilevel:
        sheet.append(["Номенклатура", "Артикул", month_label, None, None, None, None])
        sheet.append([None, None, "Количество", "Выручка", "Себестоимость", "Валовая прибыль", profitability_label])
    else:
        sheet.append([
            "Номенклатура", "Артикул", f"{month_label} Количество", f"{month_label} Выручка",
            f"{month_label} Себестоимость", f"{month_label} Валовая прибыль", f"{month_label} {profitability_label}",
        ])
    for row in rows or [["Товар A", "A-1", "1 234,56", "10 000,00", "7 000,00", "3 000,00", "30%"]]:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def vertical_xlsx_bytes(*, blocks=None, include_total=True, indent=2):
    blocks = blocks or [
        ("дек. 2025", [
            ["Тестовый товар A", 10000, 7000, 3000, 30],
            ["Тестовый товар B", 5000, None, 5000, 100],
        ]),
        ("янв. 2026", [
            ["Тестовый товар A", 12000, 8000, 4000, "33,3333%"],
            ["Тестовый товар C", 6000, 4500, 1500, 25],
        ]),
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Валовая прибыль"
    sheet.append(["Месяц", "Выручка, ₽", "Себестоимость, ₽", "Валовая прибыль, ₽", "Рентабельность"])
    sheet.append(["Номенклатура"])
    for month, rows in blocks:
        sheet.append([month])
        for row in rows:
            sheet.append(row)
            if indent is not None:
                sheet.cell(sheet.max_row, 1).alignment = Alignment(indent=indent)
    if include_total:
        sheet.append(["Итого", 33000, 19500, 13500, None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def vertical_hierarchy_xlsx(
    *, parent_indent=None, child_indent=None, include_total=True, total=(100, 70, 30),
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Месяц", "Выручка, ₽", "Себестоимость, ₽", "Валовая прибыль, ₽", "Рентабельность"])
    sheet.append(["дек. 2025"])
    sheet.append(["Тестовая группа", 100, 70, 30, 30])
    sheet.append(["Тестовый товар", 100, 70, 30, 30])
    if parent_indent is not None:
        sheet.cell(3, 1).alignment = Alignment(indent=parent_indent)
    if child_indent is not None:
        sheet.cell(4, 1).alignment = Alignment(indent=child_indent)
    if include_total:
        sheet.append(["Итого", *total, None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def vertical_counterexample_xlsx(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Месяц", "Выручка, ₽", "Себестоимость, ₽", "Валовая прибыль, ₽", "Рентабельность"])
    sheet.append(["дек. 2025"])
    for row in rows:
        sheet.append(row)
    sheet.append(["Итого", 100, 70, 30, None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def vertical_counterexample_cases():
    return {
        "parent_instead_of_children": [
            ["Тестовая группа", 100, 70, 30, 30],
            ["Тестовый товар A", "=60", "=42", "=18", "=30"],
            ["Тестовый товар B", "=40", "=28", "=12", "=30"],
        ],
        "aggregate_and_partial_children": [
            ["Тестовая группа", 60, 42, 18, 30],
            ["Тестовый товар A", 40, 28, 12, 30],
            ["Тестовый товар B", "=60", "=42", "=18", "=30"],
        ],
        "parent_with_empty_children": [
            ["Тестовая категория", 100, 70, 30, 30],
            ["Тестовый товар A", None, None, None, None],
            ["Тестовый товар B", None, None, None, None],
        ],
        "multiple_aggregates": [
            ["Тестовая категория A", 60, 42, 18, 30],
            ["Тестовая категория B", 40, 28, 12, 30],
            ["Тестовый товар A", "=60", "=42", "=18", "=30"],
            ["Тестовый товар B", "=40", "=28", "=12", "=30"],
        ],
    }


def vertical_flat_xlsx(*, include_total=True, total=(100, 70, 30)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Месяц", "Артикул", "Выручка, ₽", "Себестоимость, ₽", "Валовая прибыль, ₽", "Рентабельность"])
    sheet.append(["дек. 2025"])
    sheet.append(["Тестовый товар A", "A-1", 60, 42, 18, 30])
    sheet.append(["Тестовый товар B", "B-1", 40, 28, 12, 30])
    if include_total:
        sheet.append(["Итого", None, *total, None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def vertical_article_type_xlsx(*, include_parent=True, articles=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "Месяц", None, None, "Количество", "Выручка", "Себестоимость",
        "Валовая прибыль", "Рентабельность",
    ])
    sheet.append(["Артикул", "Тип", "Номенклатура"])
    sheet.append(["дек. 2025", None, None, 4, 18000, 7000, 11000, None])
    if include_parent:
        sheet.append(["GROUP", "Запас", "Тестовая группа", 4, 18000, 7000, 11000, None])
    rows = [
        ["TEST-001", "Запас", "Тестовый товар A", 2, 10000, 7000, 3000, 30],
        ["", "Услуга", "Тестовая услуга", 1, 5000, None, 5000, 100],
        ["TEST-003", "Работа", "Тестовая работа", 1, 3000, None, 3000, 100],
    ]
    if articles is not None:
        for row, article in zip(rows, articles):
            row[0] = article
    for row in rows:
        sheet.append(row)
        sheet.cell(sheet.max_row, 3).alignment = Alignment(indent=2)
    sheet.append(["янв. 2026", None, None, 1, 6000, 4000, 2000, None])
    sheet.append(["TEST-001", "Запас", "Тестовый товар A", 1, 6000, 4000, 2000, "33,3333"])
    sheet.cell(sheet.max_row, 3).alignment = Alignment(indent=2)
    sheet.append(["Итого", None, None, 5, 24000, 11000, 13000, None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def upload(name="monthly-profit.xlsx", data=None, **kwargs):
    return SimpleUploadedFile(
        name, data if data is not None else xlsx_bytes(**kwargs),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class MonthlyProfitParserTests(TestCase):
    def analytics(self, rows):
        defaults = {
            "period_month": date(2026, 1, 1), "nomenclature_type": "Запас",
            "revenue": Decimal("100"), "cost": Decimal("60"),
            "gross_profit": Decimal("40"),
        }
        records = [{**defaults, **row} for row in rows]
        apply_period_weighted_goods_cost(records)
        return records

    def test_weighted_cost_uses_only_actual_goods_in_same_period(self):
        rows = self.analytics([
            {"revenue": Decimal("100000"), "cost": Decimal("60000")},
            {"revenue": Decimal("50000"), "cost": Decimal("35000")},
            {"revenue": Decimal("30000"), "cost": Decimal("0"), "gross_profit": Decimal("30000")},
            {"nomenclature_type": "Услуга", "revenue": Decimal("1000000"), "cost": Decimal("1")},
            {"nomenclature_type": "Работа", "revenue": Decimal("1000000"), "cost": Decimal("1")},
        ])
        target = rows[2]
        self.assertEqual(target["cost"], Decimal("0"))
        self.assertEqual(target["calculated_cost"], Decimal("19000.00"))
        self.assertEqual(target["analytical_gross_profit"], Decimal("11000.00"))
        self.assertEqual(target["cost_source"], OneCMonthlyProfit.COST_SOURCE_CALCULATED)
        self.assertEqual(target["cost_calculation_ratio"], Decimal("0.6333333333"))

    def test_zero_cost_and_precalculated_values_never_enter_base(self):
        rows = self.analytics([
            {"revenue": Decimal("100"), "cost": Decimal("50")},
            {"revenue": Decimal("100000"), "cost": Decimal("0"), "calculated_cost": Decimal("99999")},
            {"revenue": Decimal("20"), "cost": Decimal("0")},
        ])
        self.assertEqual(rows[1]["calculated_cost"], Decimal("50000.00"))
        self.assertEqual(rows[2]["calculated_cost"], Decimal("10.00"))

    def test_missing_base_leaves_cost_and_analytical_profit_undefined(self):
        row = self.analytics([{"cost": Decimal("0")}])[0]
        self.assertEqual(row["cost"], Decimal("0"))
        self.assertIsNone(row["calculated_cost"])
        self.assertIsNone(row["analytical_gross_profit"])
        self.assertEqual(row["cost_source"], OneCMonthlyProfit.COST_SOURCE_UNDEFINED)

    def test_periods_are_independent_and_processing_is_idempotent(self):
        february = date(2026, 2, 1)
        rows = self.analytics([
            {"revenue": Decimal("100"), "cost": Decimal("50")},
            {"revenue": Decimal("20"), "cost": Decimal("0")},
            {"period_month": february, "revenue": Decimal("100"), "cost": Decimal("80")},
            {"period_month": february, "revenue": Decimal("20"), "cost": Decimal("0")},
        ])
        self.assertEqual(rows[1]["calculated_cost"], Decimal("10.00"))
        self.assertEqual(rows[3]["calculated_cost"], Decimal("16.00"))
        snapshot = [dict(row) for row in rows]
        apply_period_weighted_goods_cost(rows)
        self.assertEqual(rows, snapshot)

    def test_positive_actual_cost_is_never_replaced(self):
        row = self.analytics([{"cost": Decimal("60"), "calculated_cost": Decimal("999")}])[0]
        self.assertEqual(row["cost"], Decimal("60"))
        self.assertIsNone(row["calculated_cost"])
        self.assertEqual(row["cost_source"], OneCMonthlyProfit.COST_SOURCE_ACTUAL)

    def test_simple_report_and_multilevel_header(self):
        result = parse_monthly_profit(BytesIO(xlsx_bytes()), filename="report.xlsx")
        self.assertEqual(len(result.records), 1)
        self.assertEqual(result.records[0]["period_month"], date(2026, 1, 1))
        self.assertEqual(result.records[0]["revenue"], Decimal("10000.00"))
        self.assertEqual(result.records[0]["nomenclature_type"], "")

    def test_russian_month_variants(self):
        for label, month in (("Янв. 2026", 1), ("сент 2026", 9), ("12.2026", 12), ("2026-04", 4)):
            with self.subTest(label=label):
                result = parse_monthly_profit(BytesIO(xlsx_bytes(month_label=label)), filename="r.xlsx")
                self.assertEqual(result.records[0]["period_month"].month, month)

    def test_decimal_formats_negative_and_percent(self):
        self.assertEqual(parse_decimal("1 234,56"), Decimal("1234.56"))
        self.assertEqual(parse_decimal("1\u00a0234,56"), Decimal("1234.56"))
        self.assertEqual(parse_decimal("1,234.56"), Decimal("1234.56"))
        self.assertEqual(parse_decimal("(1 234,56)"), Decimal("-1234.56"))
        self.assertEqual(parse_decimal("30%", percent=True), Decimal("30"))
        self.assertIsNone(parse_decimal("—"))

    def test_profit_and_profitability_columns_are_distinct(self):
        result = parse_monthly_profit(
            BytesIO(xlsx_bytes(
                profitability_label="% прибыли",
                rows=[["Товар", "A-1", 1, 10000, 7000, 3000, 30]],
            )),
            filename="r.xlsx",
        )
        row = result.records[0]
        self.assertEqual(row["gross_profit"], Decimal("3000.00"))
        self.assertEqual(row["profitability_percent"], Decimal("30.0000"))

    def test_profitability_header_variants(self):
        for label in ("% прибыли", "Процент прибыли", "Рентабельность", "Рентаб., %"):
            with self.subTest(label=label):
                result = parse_monthly_profit(
                    BytesIO(xlsx_bytes(
                        profitability_label=label,
                        rows=[["Товар", "A-1", 1, 10000, 7000, 3000, 30]],
                    )),
                    filename="r.xlsx",
                )
                row = result.records[0]
                self.assertEqual(row["gross_profit"], Decimal("3000.00"))
                self.assertEqual(row["profitability_percent"], Decimal("30.0000"))

    def test_total_row_is_skipped(self):
        rows = [["Товар", "1", 1, 10, 5, 5, 50], ["Итого", "", 1, 10, 5, 5, 50]]
        result = parse_monthly_profit(BytesIO(xlsx_bytes(rows=rows)), filename="r.xlsx")
        self.assertEqual(len(result.records), 1)
        self.assertGreaterEqual(result.rows_skipped, 1)

    def test_multiple_months_are_normalized(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Отчёт за 2026 год"])
        sheet.append(["Номенклатура", "Январь 2026", None, "Февраль 2026", None])
        sheet.append([None, "Выручка", "Валовая прибыль", "Выручка", "Валовая прибыль"])
        sheet.append(["Товар", 100, 40, 200, 70])
        output = BytesIO(); workbook.save(output)
        result = parse_monthly_profit(BytesIO(output.getvalue()), filename="r.xlsx")
        self.assertEqual([row["period_month"].month for row in result.records], [1, 2])
        self.assertEqual({row["source_row_number"] for row in result.records}, {4})

    def test_product_name_starting_with_itogo_is_not_total(self):
        result = parse_monthly_profit(
            BytesIO(xlsx_bytes(rows=[["Итого вкусный товар", "A", 1, 10, 5, 5, 50]])),
            filename="r.xlsx",
        )
        self.assertEqual(len(result.records), 1)

    def test_formula_without_cached_value_creates_warning(self):
        result = parse_monthly_profit(
            BytesIO(xlsx_bytes(rows=[["Товар", "A", 1, "=1+1", 1, 1, 50]])),
            filename="r.xlsx",
        )
        self.assertGreater(result.warnings_total, 0)
        self.assertTrue(any("формулу" in item for item in result.warnings))

    def test_source_data_is_bounded(self):
        row = ["Товар", "A", 1, 10, 5, 5, 50] + ["x" * 1000] * 80
        result = parse_monthly_profit(BytesIO(xlsx_bytes(rows=[row])), filename="r.xlsx")
        cells = result.records[0]["source_data"]["cells"]
        self.assertLessEqual(len(cells), 50)
        self.assertTrue(all(not isinstance(value, str) or len(value) <= 500 for value in cells.values()))

    def test_decimal_rounding_large_values_and_profitability(self):
        result = parse_monthly_profit(
            BytesIO(xlsx_bytes(rows=[["Товар", "A", "1.1234567", "-10.125", "-2.125", "-8", "80.12345"]])),
            filename="r.xlsx",
        )
        row = result.records[0]
        self.assertEqual(row["quantity"], Decimal("1.123457"))
        self.assertEqual(row["revenue"], Decimal("-10.13"))
        self.assertEqual(row["cost"], Decimal("-2.13"))
        self.assertEqual(row["profitability_percent"], Decimal("80.1235"))
        self.assertEqual(calculate_profitability(Decimal("1"), Decimal("3")), Decimal("33.3333"))
        self.assertIsNone(calculate_profitability(Decimal("1"), Decimal("0")))

        too_large = parse_monthly_profit(
            BytesIO(xlsx_bytes(rows=[["Товар", "A", 1, "1000000000000000000", 1, 1, 1]])),
            filename="r.xlsx",
        )
        self.assertIsNone(too_large.records[0]["revenue"])
        self.assertGreater(too_large.warnings_total, 0)

    def test_unknown_year_is_critical(self):
        result = parse_monthly_profit(
            BytesIO(xlsx_bytes(year=None, month_label="Январь")), filename="r.xlsx"
        )
        self.assertTrue(result.critical_errors)
        self.assertEqual(result.records, [])

    def test_vertical_1c_report_is_normalized_and_totals_match(self):
        result = parse_monthly_profit(BytesIO(vertical_xlsx_bytes()), filename="vertical.xlsx")
        self.assertEqual(result.metadata["layout"], "vertical_1c")
        self.assertEqual(len(result.records), 4)
        self.assertEqual(result.metadata["months"], ["2025-12-01", "2026-01-01"])
        self.assertEqual(result.metadata["month_count"], 2)
        self.assertTrue(result.metadata["totals_match"])
        self.assertEqual(result.metadata["source_totals"]["revenue"], "33000.00")
        self.assertEqual(result.metadata["calculated_totals"]["gross_profit"], "13500.00")
        second = result.records[1]
        self.assertEqual(second["period_month"], date(2025, 12, 1))
        self.assertEqual(second["source_row_number"], 5)
        self.assertEqual(second["article"], "")
        self.assertEqual(second["nomenclature_type"], "")
        self.assertIsNone(second["quantity"])
        self.assertIsNone(second["cost"])
        self.assertEqual(second["gross_profit"], Decimal("5000.00"))
        self.assertEqual(second["profitability_percent"], Decimal("100.0000"))
        self.assertEqual(second["source_data"]["detected_layout"], "vertical_1c")

    def test_vertical_months_can_be_unsorted_and_empty(self):
        blocks = [
            ("май 2026", [["Тестовый товар A", 10, 7, 3, 30]]),
            ("дек. 2025", []),
            ("февр. 2026", [["Тестовый товар B", 20, 15, 5, 25]]),
        ]
        result = parse_monthly_profit(
            BytesIO(vertical_xlsx_bytes(blocks=blocks, include_total=False)), filename="vertical.xlsx"
        )
        self.assertEqual(
            result.metadata["months"], ["2026-05-01", "2025-12-01", "2026-02-01"]
        )
        self.assertEqual([row["period_month"] for row in result.records], [date(2026, 5, 1), date(2026, 2, 1)])
        self.assertNotIn("2025-12-01", result.metadata["month_totals"])
        self.assertFalse(result.critical_errors)

    def test_vertical_skips_service_rows_rows_without_month_and_exact_totals_only(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Месяц", "Выручка, ₽", "Себестоимость, ₽", "Валовая прибыль, ₽", "Рентабельность"])
        sheet.append(["Номенклатура"])
        sheet.append(["Тестовый товар до месяца", 100, 50, 50, 50])
        sheet.append(["нояб. 2025"])
        sheet.append(["Средство Итого-Пул 1 л", 100, 120, -20, -20])
        sheet.append(["Итого", 100, 120, -20, None])
        output = BytesIO(); workbook.save(output)
        result = parse_monthly_profit(BytesIO(output.getvalue()), filename="vertical.xlsx")
        self.assertEqual(len(result.records), 1)
        self.assertEqual(result.records[0]["nomenclature"], "Средство Итого-Пул 1 л")
        self.assertEqual(result.records[0]["gross_profit"], Decimal("-20.00"))
        self.assertEqual(result.records[0]["source_row_number"], 5)

    def test_vertical_profitability_header_variants(self):
        for label in ("% прибыли", "Процент прибыли", "Рентабельность", "Рентаб., %"):
            with self.subTest(label=label):
                workbook = Workbook(); sheet = workbook.active
                sheet.append(["Месяц", "Выручка, ₽", "Себестоимость, ₽", "Валовая прибыль, ₽", label])
                sheet.append(["дек. 2025"])
                sheet.append(["Тестовый товар", 10000, 7000, 3000, 30])
                output = BytesIO(); workbook.save(output)
                result = parse_monthly_profit(BytesIO(output.getvalue()), filename="vertical.xlsx")
                self.assertEqual(result.records[0]["gross_profit"], Decimal("3000.00"))
                self.assertEqual(result.records[0]["profitability_percent"], Decimal("30.0000"))

    def test_vertical_and_horizontal_records_have_same_schema(self):
        horizontal = parse_monthly_profit(BytesIO(xlsx_bytes()), filename="horizontal.xlsx")
        vertical = parse_monthly_profit(BytesIO(vertical_xlsx_bytes()), filename="vertical.xlsx")
        self.assertEqual(set(horizontal.records[0]), set(vertical.records[0]))
        self.assertEqual(horizontal.metadata["layout"], "horizontal")
        self.assertEqual(vertical.metadata["layout"], "vertical_1c")

    def test_vertical_material_with_month_word_is_not_treated_as_month(self):
        blocks = [("дек. 2025", [["Тестовый товар мартовский", 10, 7, 3, 30]])]
        result = parse_monthly_profit(
            BytesIO(vertical_xlsx_bytes(blocks=blocks, include_total=False)), filename="vertical.xlsx"
        )
        self.assertEqual(len(result.records), 1)
        self.assertEqual(result.records[0]["period_month"], date(2025, 12, 1))
        self.assertFalse(result.critical_errors)

    def test_vertical_unknown_year_is_critical(self):
        blocks = [("янв.", [["Тестовый товар", 10, 7, 3, 30]])]
        result = parse_monthly_profit(
            BytesIO(vertical_xlsx_bytes(blocks=blocks, include_total=False)), filename="vertical.xlsx"
        )
        self.assertEqual(result.metadata["layout"], "vertical_1c")
        self.assertTrue(result.critical_errors)
        self.assertEqual(result.records, [])

    def test_vertical_indent_skips_parent_aggregate_and_warns_on_total_mismatch(self):
        workbook = Workbook(); sheet = workbook.active
        sheet.append(["Месяц", "Выручка, ₽", "Себестоимость, ₽", "Валовая прибыль, ₽", "Рентабельность"])
        sheet.append(["дек. 2025"])
        sheet.append(["Тестовая группа", 100, 70, 30, 30])
        sheet.append(["Тестовый товар", 100, 70, 30, 30])
        sheet.cell(4, 1).alignment = Alignment(indent=2)
        sheet.append(["Итого", 101, 70, 31, None])
        output = BytesIO(); workbook.save(output)
        result = parse_monthly_profit(BytesIO(output.getvalue()), filename="vertical.xlsx")
        self.assertEqual(len(result.records), 1)
        self.assertEqual(result.records[0]["source_row_number"], 4)
        self.assertFalse(result.metadata["totals_match"])
        self.assertGreaterEqual(result.warnings_total, 2)
        self.assertTrue(result.critical_errors)

    def test_vertical_parent_and_child_without_indent_are_critical(self):
        result = parse_monthly_profit(
            BytesIO(vertical_hierarchy_xlsx()), filename="vertical.xlsx"
        )
        self.assertEqual(result.metadata["hierarchy_status"], "ambiguous")
        self.assertFalse(result.metadata["totals_match"])
        self.assertTrue(result.critical_errors)

    def test_vertical_parent_and_child_with_same_indent_are_critical(self):
        result = parse_monthly_profit(
            BytesIO(vertical_hierarchy_xlsx(parent_indent=2, child_indent=2)),
            filename="vertical.xlsx",
        )
        self.assertEqual(result.metadata["hierarchy_status"], "ambiguous")
        self.assertFalse(result.metadata["totals_match"])
        self.assertTrue(result.critical_errors)

    def test_vertical_ambiguous_hierarchy_without_source_totals_is_critical(self):
        result = parse_monthly_profit(
            BytesIO(vertical_hierarchy_xlsx(include_total=False)), filename="vertical.xlsx"
        )
        self.assertIsNone(result.metadata["totals_match"])
        self.assertEqual(result.metadata["hierarchy_status"], "ambiguous")
        self.assertTrue(result.critical_errors)

    def test_vertical_reliable_indent_without_source_totals_is_allowed(self):
        result = parse_monthly_profit(
            BytesIO(vertical_xlsx_bytes(include_total=False)), filename="vertical.xlsx"
        )
        self.assertEqual(len(result.records), 4)
        self.assertEqual(result.metadata["hierarchy_status"], "reliable")
        self.assertIsNone(result.metadata["totals_match"])
        self.assertFalse(result.critical_errors)

    def test_vertical_report_with_only_empty_months_is_critical(self):
        result = parse_monthly_profit(
            BytesIO(vertical_xlsx_bytes(
                blocks=[("янв. 2026", []), ("февр. 2026", [])], include_total=False,
            )),
            filename="vertical.xlsx",
        )
        self.assertEqual(result.metadata["month_count"], 2)
        self.assertEqual(result.records, [])
        self.assertTrue(result.critical_errors)

    def test_vertical_realistic_hierarchy_with_matching_totals_is_reliable(self):
        result = parse_monthly_profit(
            BytesIO(vertical_hierarchy_xlsx(parent_indent=0, child_indent=2)),
            filename="vertical.xlsx",
        )
        self.assertEqual([row["source_row_number"] for row in result.records], [4])
        self.assertEqual(result.metadata["hierarchy_status"], "reliable")
        self.assertEqual(result.metadata["hierarchy_reason"], "stable_positive_indent")
        self.assertGreater(result.metadata["aggregate_rows_skipped"], 0)
        self.assertTrue(result.metadata["totals_match"])
        self.assertFalse(result.critical_errors)

    def test_vertical_article_type_report_preserves_fields_and_hierarchy(self):
        result = parse_monthly_profit(
            BytesIO(vertical_article_type_xlsx()), filename="vertical.xlsx"
        )
        self.assertEqual(result.metadata["layout"], "vertical_1c")
        self.assertEqual(result.metadata["header_depth"], 2)
        self.assertEqual(result.metadata["hierarchy_status"], "reliable")
        self.assertEqual(result.metadata["hierarchy_reason"], "stable_positive_indent")
        self.assertGreater(result.metadata["aggregate_rows_skipped"], 0)
        self.assertEqual(len(result.records), 4)
        self.assertTrue(result.metadata["totals_match"])
        self.assertFalse(result.critical_errors)
        first, service, work, january = result.records
        self.assertEqual(first["article"], "TEST-001")
        self.assertEqual(first["nomenclature_type"], "Запас")
        self.assertEqual(first["nomenclature"], "Тестовый товар A")
        self.assertEqual(first["period_month"], date(2025, 12, 1))
        self.assertEqual(first["quantity"], Decimal("2.000000"))
        self.assertEqual(first["revenue"], Decimal("10000.00"))
        self.assertEqual(first["cost"], Decimal("7000.00"))
        self.assertEqual(first["gross_profit"], Decimal("3000.00"))
        self.assertEqual(first["profitability_percent"], Decimal("30.0000"))
        self.assertEqual(first["source_data"]["article"], "TEST-001")
        self.assertEqual(first["source_data"]["nomenclature_type"], "Запас")
        self.assertEqual(service["article"], "")
        self.assertEqual(service["nomenclature_type"], "Услуга")
        self.assertIsNone(service["cost"])
        self.assertEqual(work["nomenclature_type"], "Работа")
        self.assertEqual(january["period_month"], date(2026, 1, 1))

    def test_vertical_article_identifiers_are_normalized_without_losing_codes(self):
        result = parse_monthly_profit(
            BytesIO(vertical_article_type_xlsx(articles=[0, 123.0, "00123"])),
            filename="vertical.xlsx",
        )
        self.assertEqual([row["article"] for row in result.records[:3]], ["0", "123", "00123"])
        self.assertEqual(result.metadata["hierarchy_status"], "reliable")
        self.assertFalse(result.critical_errors)
        spaced = parse_monthly_profit(
            BytesIO(vertical_article_type_xlsx(articles=["  CODE-1  ", "", "TEST-003"])),
            filename="vertical.xlsx",
        )
        self.assertEqual(spaced.records[0]["article"], "CODE-1")

    def test_vertical_identifier_columns_are_mapped_by_normalized_headers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Месяц", None, None, "Выручка", "Себестоимость", "Валовая прибыль", "Рентабельность"])
        sheet.append(["  Тип\nноменклатуры ", " Номенклатура ", " КОД "])
        sheet.append(["дек. 2025", None, None, 100, 70, 30, None])
        sheet.append(["Запас", "Тестовая строка", " 00123 ", 100, 70, 30, 30])
        sheet.cell(4, 2).alignment = Alignment(indent=2)
        sheet.append(["Итого", None, None, 100, 70, 30, None])
        output = BytesIO()
        workbook.save(output)
        result = parse_monthly_profit(BytesIO(output.getvalue()), filename="vertical.xlsx")
        self.assertEqual(result.metadata["hierarchy_status"], "reliable")
        self.assertEqual(result.records[0]["article"], "00123")
        self.assertEqual(result.records[0]["nomenclature_type"], "Запас")
        self.assertEqual(result.records[0]["nomenclature"], "Тестовая строка")
        self.assertFalse(result.critical_errors)

    def test_nomenclature_type_classification_is_explicit(self):
        cases = (
            ("Запас", "goods"),
            ("  запас  ", "goods"),
            ("Услуга", "service"),
            ("Работа", "service"),
            ("Новый тип", "unknown"),
            ("", "unknown"),
            (None, "unknown"),
        )
        for value, expected in cases:
            with self.subTest(value=value):
                self.assertEqual(classify_nomenclature_type(value), expected)

    def test_horizontal_optional_type_column_is_mapped(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Отчёт за 2026 год"])
        sheet.append(["Номенклатура", "Артикул", "  Тип\nноменклатуры ", "Январь 2026", None])
        sheet.append([None, None, None, "Выручка", "Валовая прибыль"])
        sheet.append(["Тестовый товар", 123.0, "  Запас  ", 100, 40])
        output = BytesIO()
        workbook.save(output)
        result = parse_monthly_profit(BytesIO(output.getvalue()), filename="horizontal.xlsx")
        self.assertEqual(result.metadata["layout"], "horizontal")
        self.assertEqual(result.records[0]["article"], "123")
        self.assertEqual(result.records[0]["nomenclature_type"], "Запас")
        self.assertFalse(result.critical_errors)

    def test_vertical_ambiguous_counterexamples_stay_critical_when_totals_match(self):
        for name, rows in vertical_counterexample_cases().items():
            with self.subTest(name=name):
                result = parse_monthly_profit(
                    BytesIO(vertical_counterexample_xlsx(rows)), filename="vertical.xlsx"
                )
                self.assertEqual(result.metadata["hierarchy_status"], "ambiguous")
                self.assertEqual(result.metadata["hierarchy_reason"], "no_reliable_detail_level")
                self.assertTrue(result.metadata["totals_match"])
                self.assertTrue(result.critical_errors)

    def test_vertical_explicit_flat_schema_is_allowed(self):
        result = parse_monthly_profit(BytesIO(vertical_flat_xlsx()), filename="vertical.xlsx")
        self.assertEqual(result.metadata["hierarchy_status"], "flat")
        self.assertEqual(result.metadata["hierarchy_reason"], "explicit_flat_schema")
        self.assertEqual(len(result.records), 2)
        self.assertTrue(result.metadata["totals_match"])
        self.assertFalse(result.critical_errors)

    def test_vertical_without_indent_or_flat_evidence_is_ambiguous(self):
        result = parse_monthly_profit(
            BytesIO(vertical_xlsx_bytes(indent=None)), filename="vertical.xlsx"
        )
        self.assertEqual(result.metadata["hierarchy_status"], "ambiguous")
        self.assertTrue(result.metadata["totals_match"])
        self.assertTrue(result.critical_errors)

    def test_vertical_hierarchy_and_totals_are_independent_checks(self):
        cases = (
            (vertical_hierarchy_xlsx(parent_indent=0, child_indent=2), "reliable", True, False),
            (vertical_hierarchy_xlsx(parent_indent=0, child_indent=2, total=(101, 70, 31)), "reliable", False, True),
            (vertical_xlsx_bytes(include_total=False), "reliable", None, False),
            (vertical_flat_xlsx(), "flat", True, False),
            (vertical_flat_xlsx(include_total=False), "flat", None, False),
            (vertical_flat_xlsx(total=(101, 70, 31)), "flat", False, True),
            (vertical_counterexample_xlsx(vertical_counterexample_cases()["parent_with_empty_children"]), "ambiguous", True, True),
            (vertical_hierarchy_xlsx(), "ambiguous", False, True),
            (vertical_hierarchy_xlsx(include_total=False), "ambiguous", None, True),
        )
        for payload, status, totals_match, has_critical in cases:
            with self.subTest(status=status, totals_match=totals_match):
                result = parse_monthly_profit(BytesIO(payload), filename="vertical.xlsx")
                self.assertEqual(result.metadata["hierarchy_status"], status)
                self.assertEqual(result.metadata["totals_match"], totals_match)
                self.assertEqual(bool(result.critical_errors), has_critical)


class MonthlyProfitFormTests(TestCase):
    def test_rejects_unsupported_file(self):
        form = MonthlyProfitUploadForm(files={"report": SimpleUploadedFile("bad.xls", b"not xlsx")})
        self.assertFalse(form.is_valid())

    def test_rejects_renamed_non_xlsx(self):
        form = MonthlyProfitUploadForm(files={"report": SimpleUploadedFile("bad.xlsx", b"<html></html>")})
        self.assertFalse(form.is_valid())

    def test_rejects_too_large_file(self):
        form = MonthlyProfitUploadForm(files={"report": SimpleUploadedFile("big.xlsx", b"0" * (15 * 1024 * 1024 + 1))})
        self.assertFalse(form.is_valid())

    def test_rejects_zip_traversal(self):
        source = BytesIO(xlsx_bytes())
        output = BytesIO(source.getvalue())
        with zipfile.ZipFile(output, "a") as archive:
            archive.writestr("../escape.txt", "x")
        form = MonthlyProfitUploadForm(files={"report": SimpleUploadedFile("bad.xlsx", output.getvalue())})
        self.assertFalse(form.is_valid())


class MonthlyProfitWorkflowTests(TestCase):
    def setUp(self):
        self.private_dir = TemporaryDirectory()
        self.override = override_settings(PRIVATE_MEDIA_ROOT=self.private_dir.name)
        self.override.enable()
        self.addCleanup(self.override.disable)
        self.addCleanup(self.private_dir.cleanup)
        self.organization = Organization.objects.create(
            name="Тестовая организация", paid_until=timezone.now() + timedelta(days=30)
        )
        self.user = User.objects.create_user("owner", password="test")
        OrganizationAccess.objects.create(user=self.user, organization=self.organization, role="owner")
        self.client.force_login(self.user)

    def create_preview(self):
        return create_monthly_profit_preview(upload(), self.organization, self.user)

    def test_preview_does_not_create_profit_rows(self):
        batch = self.create_preview()
        self.assertEqual(batch.status, OneCImportBatch.STATUS_PREVIEWED)
        self.assertEqual(OneCMonthlyProfit.objects.count(), 0)

    def test_vertical_preview_does_not_create_rows_and_confirm_is_atomic(self):
        batch = create_monthly_profit_preview(
            upload(name="vertical.xlsx", data=vertical_xlsx_bytes()), self.organization, self.user
        )
        self.assertEqual(batch.metadata["report"]["layout"], "vertical_1c")
        self.assertEqual(batch.rows_detected, 4)
        self.assertEqual(OneCMonthlyProfit.objects.count(), 0)
        confirmed = confirm_monthly_profit(batch.id, self.organization, self.user)
        self.assertEqual(confirmed.rows_imported, 4)
        self.assertEqual(OneCMonthlyProfit.objects.count(), 4)

    def test_article_type_preview_and_confirm_persist_new_fields(self):
        batch = create_monthly_profit_preview(
            upload(name="article-type.xlsx", data=vertical_article_type_xlsx()),
            self.organization,
            self.user,
        )
        self.assertEqual(OneCMonthlyProfit.objects.count(), 0)
        self.assertEqual(batch.metadata["report"]["hierarchy_status"], "reliable")
        self.assertEqual(batch.metadata["preview"][0]["article"], "TEST-001")
        self.assertEqual(batch.metadata["preview"][0]["nomenclature_type"], "Запас")
        confirmed = confirm_monthly_profit(batch.id, self.organization, self.user)
        self.assertEqual(confirmed.status, OneCImportBatch.STATUS_CONFIRMED)
        rows = list(OneCMonthlyProfit.objects.filter(import_batch=batch).order_by("source_row_number"))
        self.assertEqual(len(rows), 4)
        self.assertEqual(rows[0].article, "TEST-001")
        self.assertEqual(rows[0].nomenclature_type, "Запас")
        self.assertEqual(rows[1].article, "")
        self.assertEqual(rows[1].nomenclature_type, "Услуга")

    def test_monthly_profit_model_defaults_nomenclature_type_to_empty(self):
        batch = self.create_preview()
        row = OneCMonthlyProfit.objects.create(
            import_batch=batch,
            organization=self.organization,
            period_month=date(2026, 1, 1),
            source_row_number=999,
            nomenclature="Тестовая строка",
        )
        self.assertEqual(row.nomenclature_type, "")

    def test_vertical_critical_preview_cannot_be_confirmed(self):
        batch = create_monthly_profit_preview(
            upload(name="ambiguous.xlsx", data=vertical_hierarchy_xlsx()),
            self.organization,
            self.user,
        )
        self.assertTrue(batch.metadata["critical_errors"])
        self.assertEqual(OneCMonthlyProfit.objects.count(), 0)
        with self.assertRaises(ValidationError):
            confirm_monthly_profit(batch.id, self.organization, self.user)
        batch.refresh_from_db()
        self.assertEqual(batch.status, OneCImportBatch.STATUS_FAILED)
        self.assertEqual(OneCMonthlyProfit.objects.count(), 0)

    def test_vertical_ambiguous_matching_totals_cannot_be_confirmed(self):
        for name, rows in vertical_counterexample_cases().items():
            with self.subTest(name=name):
                batch = create_monthly_profit_preview(
                    upload(name=f"{name}.xlsx", data=vertical_counterexample_xlsx(rows)),
                    self.organization,
                    self.user,
                )
                self.assertEqual(
                    batch.metadata["report"]["hierarchy_status"], "ambiguous"
                )
                self.assertTrue(batch.metadata["report"]["totals_match"])
                self.assertTrue(batch.metadata["critical_errors"])
                with self.assertRaises(ValidationError):
                    confirm_monthly_profit(batch.id, self.organization, self.user)
                batch.refresh_from_db()
                self.assertEqual(batch.status, OneCImportBatch.STATUS_FAILED)
                self.assertFalse(
                    OneCMonthlyProfit.objects.filter(import_batch=batch).exists()
                )

    def test_vertical_cancel_deletes_source_and_creates_no_rows(self):
        batch = create_monthly_profit_preview(
            upload(name="vertical-cancel.xlsx", data=vertical_xlsx_bytes()), self.organization, self.user
        )
        storage = batch.stored_file.storage
        stored_name = batch.stored_file.name
        with self.captureOnCommitCallbacks(execute=True):
            cancel_monthly_profit(batch, self.user)
        batch.refresh_from_db()
        self.assertEqual(batch.status, OneCImportBatch.STATUS_CANCELLED)
        self.assertFalse(storage.exists(stored_name))
        self.assertEqual(OneCMonthlyProfit.objects.count(), 0)

    def test_vertical_duplicate_sha_is_rejected(self):
        payload = vertical_xlsx_bytes()
        batch = create_monthly_profit_preview(
            upload(name="vertical.xlsx", data=payload), self.organization, self.user
        )
        with self.assertRaises(DuplicateImportError) as context:
            create_monthly_profit_preview(
                upload(name="same.xlsx", data=payload), self.organization, self.user
            )
        self.assertEqual(context.exception.batch.id, batch.id)

    def test_preview_metadata_has_explicit_limits(self):
        result = ParseResult(
            records=[{"nomenclature": "Товар", "revenue": Decimal("3"), "cost": Decimal("2"), "gross_profit": Decimal("1")}] * 40,
            warnings=["x" * 500] * 80,
            warnings_total=80,
        )
        metadata = _preview_metadata(result)
        self.assertEqual(len(metadata["preview"]), 30)
        self.assertEqual(len(metadata["warnings"]), 50)
        self.assertEqual(metadata["warnings_hidden"], 30)
        self.assertTrue(all(len(item) <= 300 for item in metadata["warnings"]))
        self.assertEqual(metadata["totals"]["profitability_percent"], "33.3333")

    def test_authorized_pages_render(self):
        batch = self.create_preview()
        for url in (
            reverse("finance_onec_import_list"),
            reverse("finance_onec_monthly_profit_upload"),
            reverse("finance_onec_import_preview", args=[batch.id]),
        ):
            with self.subTest(url=url):
                self.assertEqual(self.client.get(url).status_code, 200)

    def test_detail_marks_calculated_cost_with_ratio_and_explanation(self):
        batch = self.create_preview()
        batch.status = OneCImportBatch.STATUS_CONFIRMED
        batch.save(update_fields=["status"])
        OneCMonthlyProfit.objects.create(
            import_batch=batch, organization=self.organization,
            period_month=date(2026, 1, 1), source_row_number=99,
            nomenclature="Товар с расчётной стоимостью", nomenclature_type="Товар",
            revenue=Decimal("30000"), cost=Decimal("0"), gross_profit=Decimal("30000"),
            calculated_cost=Decimal("19000"),
            cost_source=OneCMonthlyProfit.COST_SOURCE_CALCULATED,
            cost_calculation_method=OneCMonthlyProfit.COST_METHOD_PERIOD_WEIGHTED_GOODS,
            cost_calculation_ratio=Decimal("0.6333333333"),
            analytical_gross_profit=Decimal("11000"),
        )
        response = self.client.get(reverse("finance_onec_import_detail", args=[batch.id]))
        self.assertContains(response, "Расчётная · 63,3%")
        self.assertContains(response, "Исходная себестоимость в 1С равна 0")
        self.assertEqual(response.context["page_obj"][0].displayed_gross_profit, Decimal("11000.00"))

    def test_admin_and_accountant_can_access(self):
        for role in ("admin", "accountant"):
            user = User.objects.create_user(f"user-{role}", password="test")
            OrganizationAccess.objects.create(user=user, organization=self.organization, role=role)
            self.client.force_login(user)
            with self.subTest(role=role):
                self.assertEqual(self.client.get(reverse("finance_onec_import_list")).status_code, 200)

    def test_confirm_creates_rows_and_second_confirm_is_rejected(self):
        batch = self.create_preview()
        confirmed = confirm_monthly_profit(batch.id, self.organization, self.user)
        self.assertEqual(confirmed.rows_imported, 1)
        self.assertEqual(OneCMonthlyProfit.objects.count(), 1)
        with self.assertRaises(ValidationError):
            confirm_monthly_profit(batch.id, self.organization, self.user)
        self.assertEqual(OneCMonthlyProfit.objects.count(), 1)

    def test_duplicate_sha_is_rejected(self):
        payload = xlsx_bytes()
        batch = create_monthly_profit_preview(upload(data=payload), self.organization, self.user)
        with self.assertRaises(DuplicateImportError) as context:
            create_monthly_profit_preview(upload(data=payload), self.organization, self.user)
        self.assertEqual(context.exception.batch.id, batch.id)

    def test_integrity_error_race_returns_existing_batch_and_removes_orphan(self):
        payload = xlsx_bytes()
        existing = create_monthly_profit_preview(upload(data=payload), self.organization, self.user)
        storage = existing.stored_file.storage
        before = set(Path(storage.location).rglob("*.xlsx"))
        with patch("django.db.models.query.QuerySet.first", return_value=None):
            with self.assertRaises(DuplicateImportError) as context:
                create_monthly_profit_preview(upload(data=payload), self.organization, self.user)
        self.assertEqual(context.exception.batch.id, existing.id)
        self.assertEqual(set(Path(storage.location).rglob("*.xlsx")), before)

    def test_storage_failure_does_not_create_batch_or_orphan(self):
        storage = OneCImportBatch._meta.get_field("stored_file").storage
        original_save = storage.save

        def save_then_fail(name, content, max_length=None):
            original_save(name, content, max_length=max_length)
            raise OSError("storage failure")

        before = set(Path(storage.location).rglob("*.xlsx"))
        with patch.object(storage, "save", side_effect=save_then_fail):
            with self.assertRaises(OSError):
                create_monthly_profit_preview(upload(), self.organization, self.user)
        self.assertEqual(OneCImportBatch.objects.count(), 0)
        self.assertEqual(set(Path(storage.location).rglob("*.xlsx")), before)

    def test_database_save_failure_removes_orphan_file(self):
        storage = OneCImportBatch._meta.get_field("stored_file").storage
        before = set(Path(storage.location).rglob("*.xlsx"))
        with patch.object(OneCImportBatch, "save", side_effect=RuntimeError("db unavailable")):
            with self.assertRaises(RuntimeError):
                create_monthly_profit_preview(upload(), self.organization, self.user)
        self.assertEqual(OneCImportBatch.objects.count(), 0)
        self.assertEqual(set(Path(storage.location).rglob("*.xlsx")), before)

    def test_parse_failure_keeps_diagnostic_file_bound_to_failed_batch(self):
        workbook = Workbook(); workbook.active.append(["Неподдерживаемый отчёт"])
        output = BytesIO(); workbook.save(output)
        bad = SimpleUploadedFile("bad.xlsx", output.getvalue())
        with self.assertRaises(Exception):
            create_monthly_profit_preview(bad, self.organization, self.user)
        batch = OneCImportBatch.objects.get()
        self.assertEqual(batch.status, OneCImportBatch.STATUS_FAILED)
        self.assertTrue(batch.stored_file.storage.exists(batch.stored_file.name))

    def test_confirm_error_rolls_back_rows(self):
        batch = self.create_preview()
        record = parse_monthly_profit(BytesIO(xlsx_bytes()), filename="r.xlsx").records[0]
        duplicate_result = ParseResult(records=[record, dict(record)])
        with patch("pool_service.finance_imports.services.parse_monthly_profit", return_value=duplicate_result):
            with self.assertRaises(IntegrityError):
                confirm_monthly_profit(batch.id, self.organization, self.user)
        self.assertEqual(OneCMonthlyProfit.objects.count(), 0)
        batch.refresh_from_db()
        self.assertEqual(batch.status, OneCImportBatch.STATUS_FAILED)

    def test_disallowed_role_gets_forbidden(self):
        manager = User.objects.create_user("manager", password="test")
        OrganizationAccess.objects.create(user=manager, organization=self.organization, role="manager")
        self.client.force_login(manager)
        response = self.client.get(reverse("finance_onec_import_list"))
        self.assertEqual(response.status_code, 403)

    def test_anonymous_and_superuser_without_organization_are_denied(self):
        self.client.logout()
        response = self.client.get(reverse("finance_onec_import_list"))
        self.assertEqual(response.status_code, 302)
        superuser = User.objects.create_superuser("root", "root@example.test", "test")
        self.client.force_login(superuser)
        self.assertEqual(self.client.get(reverse("finance_onec_import_list")).status_code, 403)

    def test_other_organization_batch_is_hidden(self):
        batch = self.create_preview()
        other_org = Organization.objects.create(name="Другая", paid_until=timezone.now() + timedelta(days=30))
        other = User.objects.create_user("other", password="test")
        OrganizationAccess.objects.create(user=other, organization=other_org, role="owner")
        self.client.force_login(other)
        response = self.client.get(reverse("finance_onec_import_preview", args=[batch.id]))
        self.assertEqual(response.status_code, 404)

    def test_tampered_batch_uuid_is_not_found(self):
        self.assertEqual(
            self.client.get(reverse("finance_onec_import_preview", args=["00000000-0000-0000-0000-000000000000"])).status_code,
            404,
        )

    def test_confirm_and_cancel_are_post_only(self):
        batch = self.create_preview()
        self.assertEqual(self.client.get(reverse("finance_onec_import_confirm", args=[batch.id])).status_code, 405)
        self.assertEqual(self.client.get(reverse("finance_onec_import_cancel", args=[batch.id])).status_code, 405)

    def test_confirm_requires_csrf(self):
        batch = self.create_preview()
        csrf_client = Client(enforce_csrf_checks=True)
        csrf_client.force_login(self.user)
        response = csrf_client.post(reverse("finance_onec_import_confirm", args=[batch.id]))
        self.assertEqual(response.status_code, 403)

    def test_cancel_removes_private_file_and_confirmed_cannot_cancel(self):
        batch = self.create_preview()
        path = batch.stored_file.path
        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(reverse("finance_onec_import_cancel", args=[batch.id]))
        self.assertEqual(response.status_code, 302)
        batch.refresh_from_db()
        self.assertEqual(batch.status, OneCImportBatch.STATUS_CANCELLED)
        self.assertFalse(batch.stored_file.storage.exists(batch.stored_file.name))

        confirmed = create_monthly_profit_preview(upload(rows=[["Товар B", "B", 1, 2, 1, 1, 50]]), self.organization, self.user)
        confirm_monthly_profit(confirmed.id, self.organization, self.user)
        self.client.post(reverse("finance_onec_import_cancel", args=[confirmed.id]))
        confirmed.refresh_from_db()
        self.assertEqual(confirmed.status, OneCImportBatch.STATUS_CONFIRMED)

    def test_stale_preview_cannot_cancel_confirmed_batch(self):
        stale = self.create_preview()
        confirm_monthly_profit(stale.id, self.organization, self.user)
        with self.assertRaises(ValidationError):
            cancel_monthly_profit(stale, self.user)
        stale.refresh_from_db()
        self.assertEqual(stale.status, OneCImportBatch.STATUS_CONFIRMED)
        self.assertTrue(stale.stored_file.storage.exists(stale.stored_file.name))

    def test_batch_source_row_unique_constraint(self):
        batch = self.create_preview()
        kwargs = dict(
            import_batch=batch, organization=self.organization, period_month=date(2026, 1, 1),
            source_row_number=1, nomenclature="Товар",
        )
        OneCMonthlyProfit.objects.create(**kwargs)
        with self.assertRaises(IntegrityError):
                with transaction.atomic(): OneCMonthlyProfit.objects.create(**kwargs)

    def test_batch_and_organization_delete_remove_only_owned_file(self):
        batch = self.create_preview()
        storage = batch.stored_file.storage
        owned_name = batch.stored_file.name
        unrelated_name = storage.save("unrelated.txt", ContentFile(b"keep"))

        with self.captureOnCommitCallbacks(execute=True):
            batch.delete()
        self.assertFalse(storage.exists(owned_name))
        self.assertTrue(storage.exists(unrelated_name))

        second = self.create_preview()
        second_name = second.stored_file.name
        with self.captureOnCommitCallbacks(execute=True):
            self.organization.delete()
        self.assertFalse(storage.exists(second_name))
        self.assertTrue(storage.exists(unrelated_name))
        storage.delete(unrelated_name)

    def test_file_deletion_waits_for_database_commit(self):
        batch = self.create_preview()
        batch_id = batch.pk
        stored_name = batch.stored_file.name
        storage = batch.stored_file.storage

        with self.assertRaises(RuntimeError):
            with transaction.atomic():
                batch.delete()
                raise RuntimeError("rollback")

        self.assertTrue(OneCImportBatch.objects.filter(pk=batch_id).exists())
        self.assertTrue(storage.exists(stored_name))

    def test_cancel_rollback_does_not_delete_file(self):
        batch = self.create_preview()
        stored_name = batch.stored_file.name
        storage = batch.stored_file.storage

        with self.assertRaises(RuntimeError):
            with transaction.atomic():
                cancel_monthly_profit(batch, self.user)
                raise RuntimeError("rollback")

        batch.refresh_from_db()
        self.assertEqual(batch.status, OneCImportBatch.STATUS_PREVIEWED)
        self.assertTrue(storage.exists(stored_name))

    def test_guarded_delete_rejects_tampered_path(self):
        batch = self.create_preview()
        storage = batch.stored_file.storage
        unrelated_name = storage.save("unrelated.txt", ContentFile(b"keep"))
        original_name = batch.stored_file.name
        batch.stored_file.name = unrelated_name
        self.assertFalse(delete_private_batch_file(batch))
        self.assertTrue(storage.exists(unrelated_name))
        self.assertTrue(storage.exists(original_name))
        storage.delete(unrelated_name)
