from io import BytesIO
from datetime import date

from openpyxl import Workbook


def _xlsx(rows, *, merged=()):
    workbook = Workbook()
    sheet = workbook.active
    for values, indent in rows:
        sheet.append(values)
        sheet.cell(sheet.max_row, 1).alignment = sheet.cell(sheet.max_row, 1).alignment.copy(
            indent=indent
        )
    for cell_range in merged:
        sheet.merge_cells(cell_range)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def payroll_xlsx(*, accrued_delta=0):
    delta = accrued_delta
    rows = [
        (("Подразделение", "", "нач. остаток", "начислено", "выплачено", "кон. остаток"), 0),
        (("Сотрудник", "Период регистрации", "", "", "", ""), 0),
        (("Основное подразделение", "", 30, 300 + delta, 240, 90 + delta), 0),
        (("Иванов Иван Иванович", "12.2024", 5, 40, 30, 15), 2),
        (("Иванов Иван Иванович", "01.2025", 5, 60 + delta, 50, 15 + delta), 2),
        (("Петров Пётр Петрович", "12.2024", 10, 80, 60, 30), 2),
        (("Петров Пётр Петрович", "01.2025", 10, 120, 100, 30), 2),
    ]
    return _xlsx(rows, merged=("A1:B1", "C1:C2", "D1:D2", "E1:E2", "F1:F2", "A3:B3"))


def cashflow_xlsx(*, payment_delta=0):
    delta = payment_delta
    rows = [
        (("Статья", "Поступления", "Платежи", "Чистый денежный поток"), 0),
        (("Месяц", "", "", ""), 0),
        (("Документ движения", "", "", ""), 0),
        (("Заработная плата", 10, 100 + delta, -90 - delta), 0),
        (("01.2025", 10, 100 + delta, -90 - delta), 2),
        (("Платёжное поручение 1", 0, 60 + delta, -60 - delta), 4),
        (("Возврат сотрудника", 10, 0, 10), 4),
        (("Платёжное поручение 2", 0, 40, -40), 4),
        (("Внутреннее перемещение", 50, 50, 0), 0),
        (("02.2025", 50, 50, 0), 2),
        (("Перевод между счетами", 50, 50, 0), 4),
    ]
    return _xlsx(rows, merged=("B1:B3", "C1:C3", "D1:D3"))


def payroll_160_rows_xlsx():
    employee_rows = []
    for employee_index in range(17):
        month_count = 10 if employee_index < 7 else 9
        for month_index in range(month_count):
            absolute_month = 2024 * 12 + 11 + month_index
            period = date(absolute_month // 12, absolute_month % 12 + 1, 1)
            employee_rows.append((
                (f"Сотрудник {employee_index + 1:02d}", period.strftime("%m.%Y"), 1, 10, 8, 3),
                2,
            ))
    rows = [
        (("Подразделение", "", "нач. остаток", "начислено", "выплачено", "кон. остаток"), 0),
        (("Сотрудник", "Период регистрации", "", "", "", ""), 0),
        (("Основное подразделение", "", 160, 1600, 1280, 480), 0),
    ]
    rows.extend(employee_rows)
    return _xlsx(rows, merged=("A1:B1", "C1:C2", "D1:D2", "E1:E2", "F1:F2", "A3:B3"))


def payroll_unmatched_months_xlsx(*, months=12, accrued_delta=0, name="Шукшин Илья Сергеевич"):
    facts = []
    for month_index in range(months):
        absolute_month = 2025 * 12 + month_index
        period = date(absolute_month // 12, absolute_month % 12 + 1, 1)
        accrued = 10 + (accrued_delta if month_index == 0 else 0)
        facts.append(((name, period.strftime("%m.%Y"), 1, accrued, 8, accrued - 7), 2))
    total_accrued = 10 * months + accrued_delta
    total_closing = 3 * months + accrued_delta
    rows = [
        (("Подразделение", "", "нач. остаток", "начислено", "выплачено", "кон. остаток"), 0),
        (("Сотрудник", "Период регистрации", "", "", "", ""), 0),
        (("Основное подразделение", "", months, total_accrued, 8 * months, total_closing), 0),
        *facts,
    ]
    return _xlsx(rows, merged=("A1:B1", "C1:C2", "D1:D2", "E1:E2", "F1:F2", "A3:B3"))
